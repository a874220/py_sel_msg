import sys
from openpyxl import Workbook
import datetime

## EXAMPLE execution from Powershell, the script will add the .xlsx extension automatically
#PS C:\Users\abrown\Documents\Tickets> python C:\Users\abrown\Documents\Moba\home\src\py_sel_msg\py_sel_msg.py
#                                             .\230907-EM0067-E1_nic_driver_crash\collectlogs_tulcp01csr013_20230907_110713\module0\sel.json.js
#                                             .\230907-EM0067-E1_nic_driver_crash\collectlogs_tulcp01csr013_20230907_110713\module0\messages.json.js
#                                             .\230907-EM0067-E1_nic_driver_crash\sel_msg_230907_EM0067

def get_raw_data():

    sel_dict = {}
    msg_dict = {}

    if len(sys.argv) < 4:
        print("Usage: py_sel_msg <sel_js_file> <msg_js_file> <xlsx output file without xlsx extension>")
        sys.exit(1)
    else:
        print(f"Input sel: {sys.argv[1]}, Input msg: {sys.argv[2]}")
        with open(sys.argv[1], 'r') as sel:
            id = 0
            parts_list = []
            for line in sel:
                parts = line.split(':')
                if "id" in parts[0]:
                    pass
                elif "\"date\"" == parts[0].strip():
                    st = ':'.join(part for part in parts[1:])
                    sts = st.strip(",.\t\n")
                    parts_list.append(sts)
                elif "\"owner\"" == parts[0].strip():
                    st = ':'.join(part for part in parts[1:])
                    sts = st.strip(",.\t\n")
                    parts_list.append(sts)
                elif "\"name\"" == parts[0].strip():
                    st = ':'.join(part for part in parts[1:])
                    sts = st.strip(",.\t\n")
                    parts_list.append(sts)
                elif "\"desc\"" == parts[0].strip():
                    st = ':'.join(part for part in parts[1:])
                    sts = st.strip(",.\t\n")
                    parts_list.append(sts)
                elif "\"t\"" == parts[0].strip():
                    st = ':'.join(part for part in parts[1:])
                    sts = st.strip(",.\t\n")
                    parts_list.append(sts)
                elif "\"se\"" == parts[0].strip():
                    st = ':'.join(part for part in parts[1:])
                    sts = st.strip(",.\t\n")
                    parts_list.append(sts)
                    sel_dict[id] = parts_list
                    parts_list = []
                    id += 1
                    continue
                else:
                    pass

        with open(sys.argv[2], 'r') as msg:
            id = 0
            parts_list = []
            for line in msg:
                parts = line.split(':')
                if "id" in parts[0]:
                    pass
                elif "date" in parts[0]:
                    st = ':'.join(part for part in parts[1:])
                    parts_list.append(st.strip(",'\t\n"))
                elif "event_type" in parts[0]:
                    st = ':'.join(part for part in parts[1:])
                    parts_list.append(st.strip(",.\t\n"))
                elif "desc" in parts[0]:
                    st = ':'.join(part for part in parts[1:])
                    parts_list.append(st.strip(",.\t\n"))
                    msg_dict[id] = parts_list
                    parts_list = []
                    id += 1
                    continue
                else:
                    pass

    return sel_dict, msg_dict



def get_datetime(val_list):
    newt = val_list[0].strip("\"")
    try:
        return datetime.datetime.strptime(newt, '%Y-%m-%d %H:%M:%S')
    except Exception as e:
        print(f"ERROR: converting datetime: {e}")
        return datetime.datetime.strptime("1959-02-02 00:00:00", '%Y-%m-%d %H:%M:%S')


#     A      B       C          D      E           F
#  [ date, owner, sensor_name, desc, sensor_type, severity ]
#
def write_sel(row, ws, current_sel_val):
    if not isinstance(current_sel_val, list):
        print("ERROR: current_sel_val is not list")
        sys.exit(1)
    else:
        try:
            rows = str(row)
        except ValueError:
            print("ERROR: Could not convert row to string")
            sys.exit(1)
        else:
            rowcols = [col+rows for col in ["A", "B", "C", "D", "E", "F"]]
            for rc, rd in zip(rowcols, current_sel_val):
                ws[rc] = rd


#     H     I           J
#  [ date, event_type, desc]
#
def write_msg(row, ws, current_msg_val):
    if not isinstance(current_msg_val, list):
        print("ERROR: current_msg_val is not list")
        sys.exit(1)
    else:
        try:
            rows = str(row)
        except ValueError:
            print("ERROR: Could not convert row to string")
            sys.exit(1)
        else:
            rowcols = [col+rows for col in ["H", "I", "J"]]
            for rc, rd in zip(rowcols, current_msg_val):
                ws[rc] = rd



def write_hdr(row, ws):
    try:
        rows = str(row)
    except ValueError:
        print("ERROR: Could not convert row to string")
        sys.exit(1)
    else:
        rowcols = [col+rows for col in ["A", "B", "C", "D", "E", "F", "H", "I", "J"]]
        rowdata = ["DATE", "OWNER", "SENSOR_NAME", "DESCRIPTION", "SENSOR TYPE", "SEVERITY", "DATE", "EVENT TYPE", "DESCRIPTION"]
        for rc, rd in zip(rowcols, rowdata):
            ws[rc] = rd


if __name__ == "__main__":
    row = 1
    wb = Workbook()
    ws = wb.active
    write_hdr(row, ws)
    row = row + 2
    sel_dict, msg_dict = get_raw_data()
    print(f"Number of Sel entries: {len(sel_dict.items())}")
    print(f"Number of Msg entries: {len(msg_dict.items())}")
    sel_iter =  iter(sel_dict.values())
    msg_iter =  iter(msg_dict.values())
    current_sel_val = next(sel_iter)
    current_msg_val = next(msg_iter)
    print(f"Sel 0: {current_sel_val}")
    print(f"Msg 0: {current_msg_val}")
    current_sel_time = get_datetime(current_sel_val)
    current_msg_time = get_datetime(current_msg_val)
    print(f"Sel 0 dt: {current_sel_time}")
    print(f"Msg 0 dt: {current_msg_time}")
    loop_sel = True
    loop_msg = True
    while loop_sel or loop_msg:
        if loop_sel and loop_msg:
            if current_sel_time <= current_msg_time:
                while current_sel_time <= current_msg_time:
                    print(f"SEL: {current_sel_val}")
                    write_sel(row, ws, current_sel_val)
                    row = row + 1
                    try:
                        current_sel_val = next(sel_iter)
                        current_sel_time = get_datetime(current_sel_val)
                    except StopIteration:
                        loop_sel = False
                        print("No more sel values")
                        break
            else:
                while current_msg_time < current_sel_time:
                    print(f"MSG: {current_msg_val}")
                    write_msg(row, ws, current_msg_val)
                    row = row + 1
                    try:
                        current_msg_val = next(msg_iter)
                        current_msg_time = get_datetime(current_msg_val)
                    except StopIteration:
                        loop_msg = False
                        print("No more msg values")
                        break
        elif loop_sel and not loop_msg:
            while loop_sel:
                # If we get here, there should be a pending value
                print(f"SEL: {current_sel_val}")
                write_sel(row, ws, current_sel_val)
                row = row + 1
                try:
                    current_sel_val = next(sel_iter)
                    current_sel_time = get_datetime(current_sel_val)
                    print(f"SEL: {current_sel_val}")
                    write_sel(row, ws, current_sel_val)
                    row = row + 1
                except StopIteration:
                    loop_sel = False
                    print("No more sel values")
                    break
        elif not loop_sel and loop_msg:
            while loop_msg:
                # If we get here, there should be a pending value
                print(f"MSG: {current_msg_val}")
                write_msg(row, ws, current_msg_val)
                row = row + 1
                try:
                    current_msg_val = next(msg_iter)
                    current_msg_time = get_datetime(current_msg_val)
                    print(f"MSG: {current_msg_val}")
                    write_msg(row, ws, current_msg_val)
                    row = row + 1
                except StopIteration:
                    loop_msg = False
                    print("No more msg values")
                    break
        else:
            break

    fname = sys.argv[3] + ".xlsx"
    wb.save(fname)
